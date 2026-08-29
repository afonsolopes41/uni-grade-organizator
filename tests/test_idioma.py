"""A interface e o Excel falam português de Portugal ou inglês."""

import io

import openpyxl

from gradeorg.consolidate import Settings, consolidate, to_json
from gradeorg.detect import build_source
from gradeorg.excel import build_workbook
from gradeorg.i18n import CATALOG, LANGUAGES, Msg, normalize_language, tr
from gradeorg.models import RawTable
from gradeorg.session import Session

CSV = (
    "Nome;Nº Aluno;Nota Final\n"
    "Ana Maria Silva;112233;15,5\n"
    "Ana Maria Silva;112299;9\n"
).encode("utf-8")


def test_as_duas_linguas_tem_as_mesmas_chaves():
    """Uma chave só em português sairia em português no meio do inglês."""
    for lang in LANGUAGES:
        assert set(CATALOG[lang]) == set(CATALOG["pt"]), f"faltam chaves em {lang}"


def test_normalizar_a_lingua():
    assert normalize_language("pt-PT") == "pt"
    assert normalize_language("en_GB") == "en"
    assert normalize_language("") == "pt"
    assert normalize_language("klingon") == "pt"


def test_mensagem_sabe_as_duas_linguas():
    mensagem = Msg("reason.header_final", header="Nota Final")
    assert mensagem.render("pt") == "cabeçalho «Nota Final» indica nota final"
    assert mensagem.render("en") == "heading «Nota Final» says it is the final grade"
    # Tratada como texto, continua a ser a versão portuguesa.
    assert str(mensagem).startswith("cabeçalho")


def fonte(source_id="s1", titulo=None, linhas=None, filename="pauta.csv", order=1):
    return build_source(source_id=source_id, filename=filename, kind="csv",
                        table=RawTable(rows=linhas, title_lines=titulo or []),
                        file_order=order)


def test_perguntas_e_avisos_em_ingles():
    sessao = Session()
    sessao.set_language("en")
    sessao.add_file("grades.csv", CSV)

    review = sessao.review()
    assert review["language"] == "en"
    perguntas = " ".join(q["title"] + q["detail"] for q in review["questions"])
    assert "Which course" in perguntas or "season" in perguntas

    resultado = sessao.result()
    avisos = resultado["conflicts"] + resultado["warnings"]
    assert avisos
    assert any("student" in a["detail"].lower() or "name" in a["detail"].lower()
               for a in avisos)


def test_epocas_traduzidas_no_json():
    src = fonte(linhas=[["Nome", "Nota Final"], ["Ana Maria Silva", "15"]])
    dados = to_json(consolidate([src]), "en")
    assert [e["label"] for e in dados["epocas"]] == ["1st Season", "2nd Season",
                                                     "Special Season"]
    uc = dados["students"][0]["subjects"]
    assert next(iter(uc.values()))["best_epoca_label"] == "1st Season"


def test_excel_em_ingles():
    src = fonte(titulo=["Network Security"],
                linhas=[["Nome", "Nº Aluno", "Nota Final"],
                        ["Ana Maria Silva", "112233", "15"]])
    settings = Settings()
    settings.language = "en"
    livro = build_workbook(consolidate([src], settings))
    stream = io.BytesIO()
    livro.save(stream)
    stream.seek(0)
    lido = openpyxl.load_workbook(stream)

    assert "Summary" in lido.sheetnames
    assert "Detail" in lido.sheetnames
    assert "Notices" in lido.sheetnames
    resumo = lido["Summary"]
    assert resumo.cell(row=1, column=1).value.startswith("Consolidated grades")


def test_excel_em_portugues_por_omissao():
    src = fonte(linhas=[["Nome", "Nota Final"], ["Ana Maria Silva", "15"]])
    livro = build_workbook(consolidate([src]))
    assert "Resumo" in livro.sheetnames
    assert "Avisos" in livro.sheetnames


def test_a_lingua_fica_guardada_entre_arranques():
    sessao = Session()
    sessao.set_language("en")
    assert Session().language == "en"
    assert tr("epoca.epoca2", Session().language) == "2nd Season"


def test_uma_cadeira_chamada_como_uma_folha_fixa_nao_a_substitui():
    """Uma UC chamada «Resumo» não pode roubar o nome à folha do resumo."""
    src = fonte(titulo=["Resumo"], linhas=[["Nome", "Nota Final"],
                                           ["Ana Maria Silva", "15"]])
    livro = build_workbook(consolidate([src]))
    assert livro.sheetnames.count("Resumo") == 1
