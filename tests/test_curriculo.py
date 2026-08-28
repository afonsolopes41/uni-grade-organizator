"""Identidade das cadeiras, currículo e médias.

A mesma cadeira pode aparecer com nomes diferentes — a pauta do teste em
português e a da época em inglês — e é o código da UC que as junta. Com o ano e
o semestre de cada cadeira preenchidos, saem as médias por semestre, por ano e
de fim de curso.
"""

import io

import openpyxl
import pytest

from gradeorg.consolidate import (
    SPLIT, Settings, consolidate, merge_questions, resolve_subjects, to_json,
)
from gradeorg.detect import build_source, guess_component_pauta, guess_semester
from gradeorg.excel import build_workbook
from gradeorg.models import RawTable
from gradeorg.session import Session


def source(sid, filename, titulo, rows, order=1):
    return build_source(sid, filename, "pdf", RawTable(rows=rows, title_lines=titulo),
                        file_order=order)


PT = source("s1", "SGR_Epoca1_pt.pdf",
            ["ISCTE - Instituto Universitário de Lisboa",
             "03713 - SGR - Segurança e Gestão de Redes", "2º Semestre"],
            [["Número", "Nome", "Nota Final"],
             ["65074", "Sergiy Nytsulenko", "11"],
             ["70033", "Filipe Bernardino", "14"]])

EN = source("s2", "SGR_Season1_en.pdf",
            ["ISCTE - University Institute of Lisbon",
             "03713 - SGR - Network Security and Management", "2nd Semester"],
            [["Number", "Name", "Grade"],
             ["65074", "Sergiy Nytsulenko", "12"],
             ["70033", "Filipe Bernardino", "15"]], order=2)


# -- a mesma cadeira em duas línguas --------------------------------------

def test_codigo_junta_a_mesma_cadeira():
    names, merged = resolve_subjects([PT, EN], Settings())
    assert len(set(names.values())) == 1
    assert merged[0]["code"] == "03713"
    assert set(merged[0]["names"]) == {"Segurança e Gestão de Redes",
                                       "Network Security and Management"}


def test_juncao_e_confirmada_pelo_utilizador():
    perguntas = merge_questions([PT, EN], Settings())
    assert len(perguntas) == 1
    valores = {o["value"] for o in perguntas[0].options}
    assert "Segurança e Gestão de Redes" in valores
    assert SPLIT in valores


def test_utilizador_escolhe_o_nome():
    settings = Settings(subject_aliases={"codigo:03713": "Segurança e Gestão de Redes"})
    assert consolidate([PT, EN], settings)["subjects"] == ["Segurança e Gestão de Redes"]
    assert merge_questions([PT, EN], settings) == []


def test_utilizador_pode_dizer_que_sao_diferentes():
    settings = Settings(subject_aliases={"codigo:03713": SPLIT})
    assert len(consolidate([PT, EN], settings)["subjects"]) == 2


def test_cadeiras_sem_codigo_juntam_se_pelo_nome():
    a = source("s1", "a.pdf", ["Álgebra Linear - Pauta"],
               [["Nome", "Nota Final"], ["Ana Silva", "14"]])
    b = source("s2", "b.pdf", ["Álgebra Linear - Pauta"],
               [["Nome", "Nota Final"], ["Rui Costa", "12"]], order=2)
    assert consolidate([a, b])["subjects"] == ["Álgebra Linear"]


# -- semestre --------------------------------------------------------------

@pytest.mark.parametrize("texto,semestre", [
    ("2º Semestre", "2"), ("1º Semestre", "1"), ("2nd Semester", "2"),
    ("1st Semester", "1"), ("Semestre 2", "2"),
    # O caso que enganava: "1st Season" não é o 1.º semestre.
    ("2nd Semester 1st Season", "2"),
    ("2º SemestreTeste 1 (30%)", "2"),
    ("Pauta 2025/2026", None),
])
def test_guess_semester(texto, semestre):
    assert guess_semester("x.pdf", RawTable(rows=[], title_lines=[texto])).value == semestre


# -- pauta de um componente só --------------------------------------------

@pytest.mark.parametrize("titulo,etiqueta,peso", [
    ("Teste 1 (30%) - 10 de abril", "Teste 1", 30),
    ("2º SemestreTeste 1 (30%)", "Teste 1", 30),
    ("Laboratório 2 (9%)", "Laboratorio 2", 9),
    ("Exame (100%)", None, None),        # o exame todo é a nota da época
    ("Pauta de Física", None, None),
])
def test_guess_component_pauta(titulo, etiqueta, peso):
    assert guess_component_pauta(RawTable(rows=[], title_lines=[titulo])) == (etiqueta, peso)


def test_pauta_de_um_teste_nao_compete_com_a_pauta_da_epoca():
    """Era o erro: a nota do Teste 1 (3,6) competia com a nota final (11)."""
    teste = source("s1", "SGR_Teste1.pdf",
                   ["03713 - SGR - Segurança e Gestão de Redes",
                    "1ª Época", "2º SemestreTeste 1 (30%) - 10 de abril de 2026"],
                   [["Número", "Nome", "Nota"],
                    ["65074", "Sergiy Nytsulenko", "3.6"]])
    epoca = source("s2", "SGR_Epoca1.pdf",
                   ["03713 - SGR - Segurança e Gestão de Redes", "1ª Época"],
                   [["Número", "Nome", "Nota Final"],
                    ["65074", "Sergiy Nytsulenko", "11"]], order=2)

    assert teste.component_label == "Teste 1"
    assert teste.component_weight == 30
    assert all(c.kind != "final" for c in teste.grade_columns())

    resultado = consolidate([teste, epoca])
    aluno = resultado["students"][0]
    uc = next(iter(aluno["subjects"].values()))
    assert uc["best"].label == "11"
    assert not [c for c in resultado["conflicts"] if c["type"] == "nota"]
    assert "Teste 1 (30%)" in uc["epocas"]["epoca1"]["components"]


# -- médias ----------------------------------------------------------------

def com_curriculo(**curriculo):
    a = source("s1", "a.pdf", ["Álgebra Linear - Pauta"],
               [["Nome", "Nota Final"], ["Ana Silva", "16"], ["Rui Costa", "8"]])
    b = source("s2", "b.pdf", ["Física - Pauta"],
               [["Nome", "Nota Final"], ["Ana Silva", "12"], ["Rui Costa", "14"]], order=2)
    c = source("s3", "c.pdf", ["Química - Pauta"],
               [["Nome", "Nota Final"], ["Ana Silva", "14"]], order=3)
    return [a, b, c], Settings(subject_curriculum=curriculo)


def medias_de(resultado, nome):
    return next(s for s in resultado["students"] if s["name"] == nome)["averages"]


def test_medias_por_semestre_e_por_ano():
    fontes, settings = com_curriculo(
        **{"Álgebra Linear": {"year": 1, "semester": 1},
           "Física": {"year": 1, "semester": 2},
           "Química": {"year": 2, "semester": 1}})
    ana = medias_de(consolidate(fontes, settings), "Ana Silva")

    por_semestre = {(s["year"], s["semester"]): s["value"] for s in ana["semesters"]}
    assert por_semestre == {(1, 1): 16.0, (1, 2): 12.0, (2, 1): 14.0}
    assert {a["year"]: a["value"] for a in ana["years"]} == {1: 14.0, 2: 14.0}
    assert ana["final"]["value"] == 14.0
    assert ana["final"]["rounded"] == 14
    assert ana["final"]["weighted"] is False


def test_media_ponderada_por_ects():
    fontes, settings = com_curriculo(
        **{"Álgebra Linear": {"year": 1, "semester": 1, "ects": 9},
           "Física": {"year": 1, "semester": 1, "ects": 3},
           "Química": {"year": 1, "semester": 1, "ects": 3}})
    ana = medias_de(consolidate(fontes, settings), "Ana Silva")
    # (16*9 + 12*3 + 14*3) / 15 = 14.8
    assert ana["final"]["value"] == 14.8
    assert ana["final"]["weighted"] is True
    assert ana["final"]["ects"] == 15


def test_so_contam_as_cadeiras_aprovadas():
    fontes, settings = com_curriculo(
        **{"Álgebra Linear": {"year": 1, "semester": 1},
           "Física": {"year": 1, "semester": 1}})
    rui = medias_de(consolidate(fontes, settings), "Rui Costa")
    assert rui["final"]["count"] == 1        # o 8 a Álgebra fica de fora
    assert rui["final"]["value"] == 14.0


def test_cadeira_sem_ano_fica_assinalada():
    fontes, settings = com_curriculo(**{"Álgebra Linear": {"year": 1, "semester": 1}})
    ana = medias_de(consolidate(fontes, settings), "Ana Silva")
    assert ana["missing_curriculum"] == ["Física", "Química"]
    assert [s["value"] for s in ana["semesters"]] == [16.0]
    assert ana["final"]["count"] == 3        # a média de curso conta-as na mesma


def test_curriculo_chega_ao_json():
    fontes, settings = com_curriculo(**{"Física": {"year": 2, "semester": 1, "ects": 6}})
    data = to_json(consolidate(fontes, settings))
    assert data["curriculum"]["Física"] == {"year": 2, "semester": 1, "ects": 6.0}
    assert data["students"][0]["averages"]["final"] is not None


# -- definições que não se apagam umas às outras ---------------------------

def test_guardar_uma_cadeira_de_cada_vez_nao_apaga_as_outras():
    """A interface guarda campo a campo; uma substituição simples perdia tudo."""
    session = Session()
    session.add_file("a.csv", b"Nome;Nota Final\nAna Silva;14\n")
    session.update(settings={"subject_curriculum": {"A": {"year": "1"}}})
    session.update(settings={"subject_curriculum": {"A": {"semester": "2"}}})
    session.update(settings={"subject_curriculum": {"B": {"year": "3"}}})
    session.update(settings={"subject_pass_marks": {"A": 10}})

    assert session.settings.subject_curriculum == {"A": {"year": 1, "semester": 2},
                                                   "B": {"year": 3}}
    assert session.settings.subject_pass_marks == {"A": 10.0}


# -- folha de médias no Excel ---------------------------------------------

def test_folha_de_medias_com_formulas():
    fontes, settings = com_curriculo(
        **{"Álgebra Linear": {"year": 1, "semester": 1, "ects": 6},
           "Física": {"year": 1, "semester": 2, "ects": 6},
           "Química": {"year": 2, "semester": 1, "ects": 6}})
    livro = build_workbook(consolidate(fontes, settings), ["a.pdf"])
    assert "Médias" in livro.sheetnames

    buffer = io.BytesIO()
    livro.save(buffer)
    buffer.seek(0)
    sheet = openpyxl.load_workbook(buffer)["Médias"]

    cabecalho = next(r for r in range(1, 12)
                     if sheet.cell(row=r, column=1).value == "Nº Aluno")
    titulos = [c.value for c in sheet[cabecalho]]
    assert "1.º ano · 1.º sem." in titulos
    assert "Média do 1.º ano" in titulos
    assert "Média de curso" in titulos

    # As médias são fórmulas: corrigir uma nota na folha da UC actualiza-as.
    coluna = titulos.index("Média de curso") + 1
    formula = sheet.cell(row=cabecalho + 1, column=coluna).value
    assert isinstance(formula, str) and formula.startswith("=IF(")


def test_pauta_de_componente_nao_pergunta_qual_e_a_nota_final():
    """A pergunta punha de volta como nota final o que era só o Teste 1."""
    from gradeorg.detect import build_questions

    teste = source("s1", "SGR_Teste1.pdf",
                   ["03713 - SGR - Segurança e Gestão de Redes",
                    "1ª Época", "Teste 1 (30%) - 10 de abril de 2026"],
                   [["Número", "Nome", "Nota"],
                    ["65074", "Sergiy Nytsulenko", "3.6"],
                    ["70033", "Filipe Bernardino", "9.5"]])
    assert teste.component_label == "Teste 1"
    assert not [q for q in build_questions([teste]) if q.type == "final_column"]


def test_semestre_da_pauta_conta_sem_o_utilizador_repetir():
    """O semestre está escrito na pauta: não faz sentido pedi-lo outra vez."""
    from gradeorg.consolidate import detected_semesters, effective_curriculum

    fonte = source("s1", "SGR.pdf",
                   ["03713 - SGR - Segurança e Gestão de Redes", "2º Semestre"],
                   [["Número", "Nome", "Nota Final"],
                    ["65074", "Sergiy Nytsulenko", "14"]])
    settings = Settings(subject_curriculum={"Segurança e Gestão de Redes": {"year": 3}})
    detected = detected_semesters([fonte], {"s1": "Segurança e Gestão de Redes"})
    assert detected == {"Segurança e Gestão de Redes": 2}

    meta = effective_curriculum("Segurança e Gestão de Redes", settings, detected)
    assert meta == {"year": 3, "semester": 2}

    ana = medias_de(consolidate([fonte], settings), "Sergiy Nytsulenko")
    assert [(s["year"], s["semester"]) for s in ana["semesters"]] == [(3, 2)]


def test_o_utilizador_manda_sobre_o_semestre_detectado():
    from gradeorg.consolidate import effective_curriculum

    settings = Settings(subject_curriculum={"Física": {"year": 1, "semester": 1}})
    meta = effective_curriculum("Física", settings, {"Física": 2})
    assert meta["semester"] == 1
