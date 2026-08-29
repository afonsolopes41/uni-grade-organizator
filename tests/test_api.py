"""Percurso completo pela API, como a página faz."""

import io

import openpyxl
import pytest

from gradeorg.app import create_app
from gradeorg.session import SESSION


@pytest.fixture
def client():
    SESSION.reset()
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client
    SESSION.reset()


CSV_PSTE = (
    "Nome;Nº Aluno;Teste 1;Nota Final;Teste 2;Nota Final 2\n"
    "Ana Maria Silva;112233;13,25;15,5;-;-\n"
    "Rui Costa Lopes;112234;5,1;RE;11,75;14\n"
).encode("utf-8")

CSV_ALG = (
    "Número;Nome;Ex 1;Ex 2;Total\n"
    "112233;Ana Maria Silva;1;2;16\n"
    "112235;Ines Santos Dias;0,5;1;11\n"
).encode("utf-8")


def upload(client, name, data):
    return client.post("/api/upload", data={"files": (io.BytesIO(data), name)},
                       content_type="multipart/form-data")


def test_pagina_inicial(client):
    response = client.get("/")
    assert response.status_code == 200
    assert b"Organizador de Notas" in response.data


def test_estado_vazio(client):
    data = client.get("/api/state").get_json()
    assert data["files"] == [] and data["sources"] == []


def test_formato_nao_suportado_e_recusado_com_explicacao(client):
    response = upload(client, "notas.docx", b"nada")
    body = response.get_json()
    assert body["accepted"] == []
    assert "não suportado" in body["rejected"][0]["error"]


def test_ficheiro_sem_tabela_e_recusado(client):
    body = upload(client, "vazio.txt", b"linha solta\n").get_json()
    assert body["accepted"] == []
    assert body["rejected"]


def test_percurso_completo(client):
    assert upload(client, "PSTe.csv", CSV_PSTE).get_json()["accepted"] == ["PSTe.csv"]
    body = upload(client, "notas.csv", CSV_ALG).get_json()
    assert body["accepted"] == ["notas.csv"]

    # A UC de "notas.csv" nao se percebe pelo nome: tem de ser perguntada.
    perguntas = {q["id"]: q for q in body["questions"]}
    assert any(q["type"] == "subject" for q in perguntas.values())

    # Aceita os palpites e escreve a UC que a aplicação não conseguiu adivinhar.
    answers = {q["id"]: q["default"] for q in perguntas.values() if q["default"]}
    sem_palpite = next(q for q in perguntas.values()
                       if q["type"] == "subject" and not q["default"])
    answers[sem_palpite["id"]] = "Álgebra Linear"
    state = client.post("/api/answers", json={"answers": answers}).get_json()
    assert state["questions"] == []

    results = client.get("/api/results").get_json()
    assert results["stats"]["students"] == 3
    assert "Álgebra Linear" in results["subjects"]

    ana = next(s for s in results["students"] if s["name"] == "Ana Maria Silva")
    assert len(ana["subjects"]) == 2

    rui = next(s for s in results["students"] if s["name"] == "Rui Costa Lopes")
    pste = rui["subjects"]["PSTe"]
    assert pste["best"]["label"] == "14"
    assert pste["best_epoca"] == "epoca2"

    excel = client.post("/api/export", json={})
    assert excel.status_code == 200
    book = openpyxl.load_workbook(io.BytesIO(excel.data))
    assert "Resumo" in book.sheetnames and "Álgebra Linear" in book.sheetnames


def test_export_apenas_de_alguns_alunos(client):
    upload(client, "PSTe.csv", CSV_PSTE)
    response = client.post("/api/export", json={"students": ["id:112234"]})
    book = openpyxl.load_workbook(io.BytesIO(response.data))
    nomes = [book["Resumo"].cell(row=r, column=2).value for r in range(7, 10)]
    assert "Rui Costa Lopes" in nomes
    assert "Ana Maria Silva" not in nomes


def test_export_sem_ficheiros_da_erro_util(client):
    response = client.post("/api/export", json={})
    assert response.status_code == 400
    assert "nenhum ficheiro" in response.get_json()["error"]


def test_override_de_coluna_muda_o_resultado(client):
    upload(client, "PSTe.csv", CSV_PSTE)
    source_id = client.get("/api/state").get_json()["sources"][0]["id"]

    # "Nota Final" (col. 3) passa a ser a nota final da 1.a epoca, em vez de nada.
    client.post("/api/answers", json={
        "overrides": {source_id: {"3": {"role": "grade", "epoca": "epoca1", "kind": "final"}}}})
    results = client.get("/api/results").get_json()
    ana = next(s for s in results["students"] if s["name"] == "Ana Maria Silva")
    assert ana["subjects"]["PSTe"]["epocas"]["epoca1"]["grade"]["label"] == "15,5"


def test_nota_minima_muda_as_aprovacoes(client):
    upload(client, "PSTe.csv", CSV_PSTE)
    antes = client.get("/api/results").get_json()["stats"]["approved"]
    client.post("/api/answers", json={"settings": {"pass_mark": 16}})
    depois = client.get("/api/results").get_json()["stats"]["approved"]
    assert depois < antes


def test_remover_ficheiro(client):
    upload(client, "PSTe.csv", CSV_PSTE)
    state = client.post("/api/files/remove", json={"name": "PSTe.csv"}).get_json()
    assert state["files"] == []


def test_reset(client):
    upload(client, "PSTe.csv", CSV_PSTE)
    assert client.post("/api/reset").get_json()["files"] == []


def test_retirar_uma_resposta_repoe_a_deteccao(client):
    upload(client, "notas.csv", CSV_ALG)
    perguntas = client.get("/api/state").get_json()["questions"]
    subject = next(q["id"] for q in perguntas if q["type"] == "subject")
    answers = {q["id"]: (q["default"] or "Álgebra Linear") for q in perguntas}
    client.post("/api/answers", json={"answers": answers})
    assert client.get("/api/state").get_json()["questions"] == []

    client.post("/api/answers", json={"answers": {subject: ""}})
    restantes = client.get("/api/state").get_json()["questions"]
    assert [q["id"] for q in restantes] == [subject]


# -- gerir cadeiras e língua pela API --------------------------------------

def test_api_muda_o_nome_de_uma_cadeira(client):
    upload(client, "pauta.csv", CSV_PSTE)
    antes = client.get("/api/state").get_json()["subjects"][0]

    resposta = client.post("/api/subjects", json={
        "action": "rename", "subject": antes, "name": "Física Geral"})
    assert resposta.status_code == 200
    assert "Física Geral" in resposta.get_json()["subjects"]


def test_api_apaga_e_repoe_uma_cadeira(client):
    upload(client, "pauta.csv", CSV_PSTE)
    upload(client, "algebra.csv", CSV_ALG)
    subjects = client.get("/api/state").get_json()["subjects"]
    alvo = subjects[0]

    depois = client.post("/api/subjects", json={"action": "remove", "subject": alvo})
    assert depois.get_json()["removed_subjects"] == [alvo]
    assert alvo not in client.get("/api/results").get_json()["subjects"]

    client.post("/api/subjects", json={"action": "restore", "subject": alvo})
    assert alvo in client.get("/api/results").get_json()["subjects"]


def test_api_recusa_uma_accao_desconhecida(client):
    upload(client, "pauta.csv", CSV_PSTE)
    assert client.post("/api/subjects", json={"action": "voar"}).status_code == 400


def test_api_troca_de_lingua(client):
    upload(client, "pauta.csv", CSV_PSTE)
    resposta = client.post("/api/language", json={"language": "en"})
    assert resposta.get_json()["language"] == "en"

    resultados = client.get("/api/results").get_json()
    assert [e["label"] for e in resultados["epocas"]][0] == "1st Season"

    excel = client.post("/api/export", json={})
    livro = openpyxl.load_workbook(io.BytesIO(excel.data))
    assert "Summary" in livro.sheetnames


def test_api_diz_de_que_ficheiros_vem_cada_cadeira(client):
    upload(client, "pauta.csv", CSV_PSTE)
    ficheiros = client.get("/api/state").get_json()["subject_files"]
    assert [*ficheiros.values()][0] == ["pauta.csv"]


def test_api_cria_uma_cadeira_e_aponta_lhe_um_ficheiro(client):
    upload(client, "pauta.csv", CSV_PSTE)
    criada = client.post("/api/subjects", json={"action": "create", "name": "Óptica"})
    assert "Óptica" in criada.get_json()["subjects"]

    apontada = client.post("/api/subjects", json={
        "action": "assign", "file": "pauta.csv", "subject": "Óptica"})
    assert apontada.get_json()["subject_files"]["Óptica"] == ["pauta.csv"]


def test_api_abre_o_documento_de_uma_fonte(client):
    upload(client, "pauta.csv", CSV_PSTE)
    fonte = client.get("/api/state").get_json()["sources"][0]["id"]
    resposta = client.get(f"/api/document/{fonte}")
    assert resposta.status_code == 200
    assert b"Ana Maria Silva" in resposta.data
    assert client.get("/api/document/f99s0").status_code == 404


def test_api_confirma_uma_pauta(client):
    upload(client, "pauta.csv", CSV_PSTE)
    fonte = client.get("/api/state").get_json()["sources"][0]["id"]
    depois = client.post("/api/sources/confirm", json={"source_id": fonte})
    assert depois.get_json()["confirmed_sources"] == [fonte]
    voltou = client.post("/api/sources/confirm",
                         json={"source_id": fonte, "confirmed": False})
    assert voltou.get_json()["confirmed_sources"] == []


def test_api_traz_as_notas_finais_arredondadas(client):
    upload(client, "decimas.csv",
           "Nome;Nº Aluno;Nota Final\nAna Maria Silva;112233;13,4\n"
           "Rui Costa Lopes;112234;13,5\n".encode("utf-8"))
    alunos = {a["name"]: a for a in client.get("/api/results").get_json()["students"]}
    uc = next(iter(alunos["Ana Maria Silva"]["subjects"]))
    assert alunos["Ana Maria Silva"]["subjects"][uc]["best_rounded"] == 13
    assert alunos["Rui Costa Lopes"]["subjects"][uc]["best_rounded"] == 14
