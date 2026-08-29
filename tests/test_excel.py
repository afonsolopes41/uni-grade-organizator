"""O livro gerado tem as folhas, os valores e as formulas certas."""

import openpyxl
import pytest

from gradeorg.consolidate import consolidate
from gradeorg.detect import build_source
from gradeorg.excel import SUBJECT_HEADER_ROW, build_workbook
from gradeorg.models import RawTable

SAFE_FUNCTIONS = {"IF", "IFERROR", "INDEX", "MATCH", "MAX", "COUNT", "COUNTIF",
                  "COUNTIFS", "AVERAGE", "ROUND", "ISNUMBER", "NOT", "SUM", "AND"}


@pytest.fixture
def workbook(tmp_path):
    src = build_source("s1", "pauta.pdf", "pdf", RawTable(
        rows=[["Nome", "Nº Aluno", "Projeto", "Teste 1", "Nota Final",
               "Teste 2", "Nota Final 2"],
              ["Ana Maria Silva", "112233", "17", "13,25", "15,5", "-", "-"],
              ["Rui Costa Lopes", "112234", "16", "5,1", "RE", "11,75", "14"],
              ["Ines Santos Dias", "112235", "14", "0", "NA", "-", "-"]],
        title_lines=["Análise Matemática - Pauta 2025/2026"],
        footer_lines=["2026/06/25"]))
    book = build_workbook(consolidate([src]), ["pauta.pdf"])
    path = tmp_path / "saida.xlsx"
    book.save(path)
    return openpyxl.load_workbook(path)


def test_tem_as_folhas_esperadas(workbook):
    assert workbook.sheetnames == ["Resumo", "Análise Matemática", "Médias",
                                   "Detalhe", "Avisos"]


def test_folha_da_uc_tem_as_epocas_e_a_melhor_nota(workbook):
    sheet = workbook["Análise Matemática"]
    header = [sheet.cell(row=SUBJECT_HEADER_ROW, column=c).value for c in range(1, 10)]
    assert header[:9] == ["Nº Aluno", "Nome", "1.ª Época", "2.ª Época", "Época Especial",
                          "Melhor Nota", "Nota Final", "Época da melhor", "Estado"]

    linhas = {sheet.cell(row=r, column=2).value: r
              for r in range(SUBJECT_HEADER_ROW + 1, SUBJECT_HEADER_ROW + 4)}
    ana = linhas["Ana Maria Silva"]
    assert sheet.cell(row=ana, column=3).value == 15.5
    assert sheet.cell(row=ana, column=4).value == "—"

    rui = linhas["Rui Costa Lopes"]
    assert sheet.cell(row=rui, column=3).value == "Reprovado"
    assert sheet.cell(row=rui, column=4).value == 14


def test_numero_de_aluno_fica_numerico(workbook):
    sheet = workbook["Análise Matemática"]
    assert sheet.cell(row=SUBJECT_HEADER_ROW + 1, column=1).value == 112233


def test_melhor_nota_e_uma_formula_que_recalcula(workbook):
    sheet = workbook["Análise Matemática"]
    row = SUBJECT_HEADER_ROW + 1
    assert sheet.cell(row=row, column=6).value == f'=IF(COUNT(C{row}:E{row})=0,"—",MAX(C{row}:E{row}))'
    assert sheet.cell(row=row, column=7).value == f'=IF(ISNUMBER(F{row}),ROUND(F{row},0),"—")'


def resumo_header_row(sheet):
    for row in range(1, 30):
        if sheet.cell(row=row, column=1).value == "Nº Aluno":
            return row
    raise AssertionError("cabeçalho do Resumo não encontrado")


def test_resumo_puxa_a_nota_final_da_folha_da_uc(workbook):
    """Coluna G da folha da UC: a nota final, já arredondada."""
    resumo = workbook["Resumo"]
    linha = resumo_header_row(resumo) + 1
    primeira = SUBJECT_HEADER_ROW + 1
    ultima = SUBJECT_HEADER_ROW + 3
    formula = resumo.cell(row=linha, column=3).value
    assert formula.startswith(
        f"=IFERROR(INDEX('Análise Matemática'!$G${primeira}:$G${ultima}")
    assert (f"MATCH($A{linha},'Análise Matemática'!$A${primeira}:$A${ultima},0)"
            in formula)


def test_cada_uc_tem_a_sua_nota_minima_editavel(workbook):
    from gradeorg.excel import SUBJECT_PASS_ROW
    sheet = workbook["Análise Matemática"]
    assert sheet.cell(row=SUBJECT_PASS_ROW, column=1).value == "Nota mínima de aprovação"
    assert sheet.cell(row=SUBJECT_PASS_ROW, column=2).value == 9.5

    # O Estado desta folha compara com essa célula, não com um número fixo.
    estado = sheet.cell(row=SUBJECT_HEADER_ROW + 1, column=9).value
    assert f"$B${SUBJECT_PASS_ROW}" in estado

    # E o Resumo espelha-a.
    resumo = workbook["Resumo"]
    espelhos = [c.value for row in resumo.iter_rows(max_row=resumo_header_row(resumo))
                for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    assert any(f"'Análise Matemática'!$B${SUBJECT_PASS_ROW}" in e for e in espelhos)


def test_a_folha_da_uc_so_tem_a_nota_final(workbook):
    """Componentes (Projeto, Teste 1) nao geram colunas: so conta a nota final."""
    sheet = workbook["Análise Matemática"]
    headers = [c.value for c in sheet[SUBJECT_HEADER_ROW] if c.value]
    assert headers[-1] == "Origem da melhor nota"
    assert not [h for h in headers if "Projeto" in h or "Teste" in h]


def test_detalhe_tem_uma_linha_por_nota_final(workbook):
    sheet = workbook["Detalhe"]
    rows = [[c.value for c in row] for row in sheet.iter_rows(min_row=5)]
    rows = [r for r in rows if r[1]]
    assert sorted({r[4] for r in rows}) == ["Nota final"]
    ana = [r for r in rows if r[1] == "Ana Maria Silva"]
    assert {r[5] for r in ana} == {"Nota Final"}


def test_so_usa_funcoes_que_o_excel_antigo_conhece(workbook):
    import re
    used = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    used |= set(re.findall(r"([A-Z_][A-Z0-9_.]*)\s*\(", cell.value))
    assert used <= SAFE_FUNCTIONS, f"funções arriscadas: {used - SAFE_FUNCTIONS}"


def test_notas_minimas_diferentes_por_uc(tmp_path):
    from gradeorg.consolidate import Settings
    from gradeorg.excel import SUBJECT_PASS_ROW

    a = build_source("s1", "a.pdf", "pdf", RawTable(
        rows=[["Nome", "Nota Final"], ["Ana Maria Silva", "9,7"]],
        title_lines=["Análise Matemática - Pauta"]))
    b = build_source("s2", "b.pdf", "pdf", RawTable(
        rows=[["Nome", "Nota Final"], ["Ana Maria Silva", "9,7"]],
        title_lines=["Álgebra Linear - Pauta"]), file_order=2)
    settings = Settings(pass_mark=9.5, subject_pass_marks={"Análise Matemática": 12.0})

    book = build_workbook(consolidate([a, b], settings), ["a.pdf", "b.pdf"])
    assert book["Análise Matemática"].cell(row=SUBJECT_PASS_ROW, column=2).value == 12.0
    assert book["Álgebra Linear"].cell(row=SUBJECT_PASS_ROW, column=2).value == 9.5


def test_selecao_de_alunos_e_de_ucs(tmp_path):
    src = build_source("s1", "pauta.pdf", "pdf", RawTable(
        rows=[["Nome", "Nº Aluno", "Nota Final"],
              ["Ana Maria Silva", "112233", "15"],
              ["Rui Costa Lopes", "112234", "12"]],
        title_lines=["Análise Matemática - Pauta"]))
    result = consolidate([src])
    book = build_workbook(result, ["pauta.pdf"], selected_students=["id:112233"])
    sheet = book["Análise Matemática"]
    nomes = [sheet.cell(row=r, column=2).value
             for r in range(SUBJECT_HEADER_ROW + 1, SUBJECT_HEADER_ROW + 4)]
    assert "Ana Maria Silva" in nomes
    assert "Rui Costa Lopes" not in nomes


def test_uc_sem_alunos_seleccionados_nao_gera_folha():
    a = build_source("s1", "a.pdf", "pdf", RawTable(
        rows=[["Nome", "Nº Aluno", "Nota Final"], ["Ana Maria Silva", "112233", "15"]],
        title_lines=["Análise Matemática - Pauta"]))
    b = build_source("s2", "b.pdf", "pdf", RawTable(
        rows=[["Nome", "Nº Aluno", "Nota Final"], ["Rui Costa Lopes", "112234", "12"]],
        title_lines=["Álgebra Linear - Pauta"]), file_order=2)
    book = build_workbook(consolidate([a, b]), ["a.pdf", "b.pdf"],
                          selected_students=["id:112233"])
    assert "Álgebra Linear" not in book.sheetnames
    assert "Análise Matemática" in book.sheetnames


def test_livro_sem_conflitos_diz_que_esta_tudo_bem(workbook):
    avisos = workbook["Avisos"]
    textos = [c.value for row in avisos.iter_rows(max_row=10) for c in row
              if isinstance(c.value, str)]
    assert any("Sem conflitos" in t for t in textos)


# -- nota final arredondada e grupos por ano/semestre ----------------------

def _livro(rows, curriculum=None, **kwargs):
    from gradeorg.consolidate import Settings
    settings = Settings(subject_curriculum=curriculum or {})
    src = build_source("s1", "pauta.pdf", "pdf", RawTable(
        rows=rows, title_lines=["Análise Matemática - Pauta 2025/2026"]))
    return build_workbook(consolidate([src], settings), ["pauta.pdf"], **kwargs)


def test_resumo_escreve_a_nota_final_arredondada(tmp_path):
    """Sem número de aluno não há fórmula: fica o valor -- e é o arredondado."""
    book = _livro([["Nome", "Nota Final"],
                   ["Ana Maria Silva", "13,4"],
                   ["Rui Costa Lopes", "13,5"]])
    path = tmp_path / "s.xlsx"
    book.save(path)
    resumo = openpyxl.load_workbook(path)["Resumo"]
    valores = {}
    for row in resumo.iter_rows(min_row=2, values_only=True):
        if row[1] in ("Ana Maria Silva", "Rui Costa Lopes"):
            valores[row[1]] = row[2]
    assert valores == {"Ana Maria Silva": 13, "Rui Costa Lopes": 14}


def test_resumo_diz_o_ano_e_o_semestre_no_cabecalho_da_uc(tmp_path):
    book = _livro([["Nome", "Nº Aluno", "Nota Final"],
                   ["Ana Maria Silva", "112233", "15"]],
                  curriculum={"Análise Matemática": {"year": 2, "semester": 1}})
    path = tmp_path / "s.xlsx"
    book.save(path)
    resumo = openpyxl.load_workbook(path)["Resumo"]
    cabecalhos = [c.value for row in resumo.iter_rows(min_row=1, max_row=12)
                  for c in row if isinstance(c.value, str) and "Análise" in c.value]
    assert any("2.º ano · 1.º sem." in c for c in cabecalhos)
