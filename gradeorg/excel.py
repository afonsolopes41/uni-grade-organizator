"""Geracao do ficheiro Excel final.

O livro tem quatro tipos de folha:

* **Resumo** -- um aluno por linha, uma coluna por UC, com media e contagens;
* **uma folha por UC** -- as tres epocas lado a lado e a melhor nota;
* **Detalhe** -- formato longo, uma linha por nota, com a origem;
* **Avisos** -- conflitos e coisas que convem o utilizador confirmar.

As celulas calculadas (melhor nota, media, contagens) sao formulas, para que o
ficheiro continue certo se alguem corrigir uma nota a mao.
"""

from __future__ import annotations

import datetime as dt
import re
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.worksheet import Worksheet

from .i18n import (DEFAULT_LANGUAGE, epoca_label, group_label, notice_type,
                   render, tr)
from .models import EPOCAS
from .normalize import Grade, round_grade

FONT = "Arial"

INK = "1F3864"          # azul escuro do cabecalho
INK_SOFT = "2E75B6"     # azul de apoio
BAND = "F4F7FB"         # riscas alternadas
LINE = "D9E1F2"         # grelha
GREEN, GREEN_BG = "1E6B34", "E3F4E7"
RED, RED_BG = "B3261E", "FBE7E5"
AMBER, AMBER_BG = "8A5300", "FDF1DF"
MUTED = "6B7280"

THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GRADE_FORMAT = "0.##"

#: Linha do cabecalho nas folhas de UC.
#: 1 titulo, 2 subtitulo, 3 nota minima da UC, 4 em branco, 5 cabecalho.
SUBJECT_PASS_ROW = 3
SUBJECT_HEADER_ROW = 5


def _subject_extent(count: int):
    """``(primeira_linha, ultima_linha)`` dos dados de uma folha de UC."""
    first = SUBJECT_HEADER_ROW + 1
    return first, first + max(count, 1) - 1


# --------------------------------------------------------------------------
# Utilitarios de estilo
# --------------------------------------------------------------------------

def _title_block(sheet: Worksheet, width: int, title: str, subtitle: str = "") -> int:
    """Faixa de titulo no topo da folha. Devolve a proxima linha livre."""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(width, 2))
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(name=FONT, size=15, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sheet.row_dimensions[1].height = 30

    if subtitle:
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(width, 2))
        cell = sheet.cell(row=2, column=1, value=subtitle)
        cell.font = Font(name=FONT, size=9, color=MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sheet.row_dimensions[2].height = 18
        return 4
    return 3


def _header_row(sheet: Worksheet, row: int, headers: list, widths: Optional[list] = None) -> None:
    for index, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=index, value=text)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK_SOFT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    sheet.row_dimensions[row].height = 30
    if widths:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width


def _style_body(sheet: Worksheet, first_row: int, last_row: int, width: int,
                centre_from: int = 3) -> None:
    band = PatternFill("solid", fgColor=BAND)
    for row in range(first_row, last_row + 1):
        striped = (row - first_row) % 2 == 1
        for column in range(1, width + 1):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if striped:
                cell.fill = band
            if column >= centre_from:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _grade_rules(sheet: Worksheet, ref: str, pass_mark: float) -> None:
    """Vermelho abaixo da nota minima, verde a partir de 16."""
    sheet.conditional_formatting.add(ref, CellIsRule(
        operator="lessThan", formula=[str(pass_mark)],
        font=Font(name=FONT, size=10, bold=True, color=RED),
        fill=PatternFill("solid", bgColor=RED_BG)))
    sheet.conditional_formatting.add(ref, CellIsRule(
        operator="greaterThanOrEqual", formula=["16"],
        font=Font(name=FONT, size=10, bold=True, color=GREEN),
        fill=PatternFill("solid", bgColor=GREEN_BG)))


_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(base: str, used: set) -> str:
    """Nome de folha valido (<=31 caracteres, sem simbolos proibidos, unico)."""
    name = _INVALID_SHEET.sub("-", base).strip() or "UC"
    name = name[:31].strip()
    candidate, counter = name, 2
    while candidate.lower() in used:
        suffix = f" ({counter})"
        candidate = name[: 31 - len(suffix)].strip() + suffix
        counter += 1
    used.add(candidate.lower())
    return candidate


def _grade_cell(sheet: Worksheet, row: int, column: int, grade: Optional[Grade],
                lang: str = DEFAULT_LANGUAGE):
    """Escreve uma nota: numero quando ha numero, texto do estado caso contrario."""
    cell = sheet.cell(row=row, column=column)
    if grade is None:
        cell.value = "—"
        cell.font = Font(name=FONT, size=10, color=MUTED)
        return cell
    if grade.value is not None:
        cell.value = round(grade.value, 4)
        cell.number_format = GRADE_FORMAT
    else:
        cell.value = grade.label_in(lang)
        cell.font = Font(name=FONT, size=10, italic=True, color=MUTED)
    return cell


def _id_cell(sheet: Worksheet, row: int, column: int, student_id: Optional[str]):
    cell = sheet.cell(row=row, column=column)
    if student_id and student_id.isdigit():
        cell.value = int(student_id)
        cell.number_format = "0"
    else:
        cell.value = student_id or "—"
    return cell


# --------------------------------------------------------------------------
# Livro completo
# --------------------------------------------------------------------------

def build_workbook(result: dict, source_labels: Optional[list] = None,
                   selected_students: Optional[list] = None,
                   selected_subjects: Optional[list] = None,
                   lang: Optional[str] = None) -> Workbook:
    """Constroi o livro a partir do resultado de ``consolidate.consolidate``."""
    settings = result["settings"]
    lang = lang or settings.get("language") or DEFAULT_LANGUAGE
    pass_mark = float(settings.get("pass_mark", 9.5))
    pass_marks = {k: float(v) for k, v in (result.get("pass_marks") or {}).items()}

    students = result["students"]
    if selected_students:
        # A ordem da seleccao manda: e a que o utilizador poe na tabela,
        # arrastando os alunos, e a folha tem de sair como a tabela esta.
        ordem = {chave: i for i, chave in enumerate(selected_students)}
        fim = len(ordem)

        def lugar(student: dict) -> int:
            posicao = ordem.get(student["key"])
            if posicao is None:
                posicao = ordem.get(student["name"])
            return fim if posicao is None else posicao

        students = [s for s in students
                    if s["key"] in ordem or s["name"] in ordem]
        students.sort(key=lugar)
    subjects = list(result["subjects"])
    if selected_subjects:
        subjects = [s for s in subjects if s in set(selected_subjects)]
    # So entram UCs com alunos na selecao: uma folha vazia deixaria as formulas
    # do Resumo a apontar para o nada.
    subjects = [s for s in subjects if any(s in st["subjects"] for st in students)]

    workbook = Workbook()
    workbook.remove(workbook.active)

    stamp = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    origem = ", ".join(source_labels or []) or "—"

    used_names: set = {tr(key, lang).lower() for key in
                       ("xl.sheet.summary", "xl.sheet.averages",
                        "xl.sheet.detail", "xl.sheet.notices")}
    subject_sheets: dict = {}
    subject_extent: dict = {}
    for subject in subjects:
        subject_sheets[subject] = _sheet_name(subject, used_names)
        count = sum(1 for s in students if subject in s["subjects"])
        subject_extent[subject] = _subject_extent(count)

    files_by_subject = result.get("subject_files") or {}

    summary_name = tr("xl.sheet.summary", lang)
    summary = workbook.create_sheet(summary_name)
    curriculum = result.get("curriculum") or {}
    summary_first_row = _build_summary(summary, students, subjects, subject_sheets,
                                       subject_extent, pass_marks, pass_mark,
                                       stamp, origem, lang, curriculum)

    for subject in subjects:
        sheet = workbook.create_sheet(subject_sheets[subject])
        _build_subject_sheet(sheet, subject, students,
                             pass_marks.get(subject, pass_mark), stamp,
                             files_by_subject.get(subject) or [], lang)

    if subjects:
        _build_averages(workbook.create_sheet(tr("xl.sheet.averages", lang)),
                        students, subjects, subject_sheets, summary_first_row,
                        pass_mark, curriculum, stamp, summary_name, lang)

    _build_detail(workbook.create_sheet(tr("xl.sheet.detail", lang)),
                  students, subjects, stamp, lang)
    _build_notices(workbook.create_sheet(tr("xl.sheet.notices", lang)),
                   result, stamp, lang)

    workbook.active = 0
    return workbook


# --------------------------------------------------------------------------
# Folha "Resumo"
# --------------------------------------------------------------------------

def _build_summary(sheet: Worksheet, students: list, subjects: list,
                   subject_sheets: dict, subject_extent: dict,
                   pass_marks: dict, default_pass: float, stamp: str, origem: str,
                   lang: str = DEFAULT_LANGUAGE,
                   curriculum: Optional[dict] = None) -> int:
    width = max(2 + len(subjects) + 3, 4)
    row = _title_block(sheet, width, tr("xl.summary.title", lang),
                       tr("xl.summary.subtitle", lang, stamp=stamp, files=origem))

    # A nota minima de cada cadeira vive na folha dessa cadeira; aqui fica um
    # espelho, para se ver tudo de uma vez.
    sheet.cell(row=row, column=1, value=tr("xl.summary.pass_row", lang)).font = Font(
        name=FONT, size=11, bold=True, color=INK)
    sheet.cell(row=row, column=3,
               value=tr("xl.summary.pass_hint", lang)).font = Font(
        name=FONT, size=9, italic=True, color=MUTED)
    row += 1

    pass_refs = {}
    for subject in subjects:
        target = quote_sheetname(subject_sheets[subject])
        sheet.cell(row=row, column=1, value=subject).font = Font(name=FONT, size=10)
        mirror = sheet.cell(row=row, column=2, value=f"={target}!$B${SUBJECT_PASS_ROW}")
        mirror.number_format = GRADE_FORMAT
        mirror.font = Font(name=FONT, size=10, bold=True, color=GREEN)
        mirror.alignment = Alignment(horizontal="center")
        pass_refs[subject] = f"{quote_sheetname(sheet.title)}!$B${row}"
        row += 1
    if not subjects:
        pass_refs = {}
    row += 1

    # O cabecalho de cada UC leva o ano e o semestre por baixo do nome: e assim
    # que se veem os grupos sem partir a tabela em varias.
    curriculum = curriculum or {}
    titulos = []
    for subject in subjects:
        meta = curriculum.get(subject) or {}
        if meta.get("year") is not None or meta.get("semester") is not None:
            titulos.append(f"{subject}\n"
                           + group_label(meta.get("year"), meta.get("semester"), lang))
        else:
            titulos.append(subject)
    headers = ([tr("xl.col.student_id", lang), tr("xl.col.name", lang)] + titulos
               + [tr("xl.col.average", lang), tr("xl.col.approved", lang),
                  tr("xl.col.subjects", lang)])
    widths = [11, 42] + [max(14, min(24, len(s) // 2 + 8)) for s in subjects] + [10, 11, 8]
    header_row = row
    _header_row(sheet, header_row, headers, widths)

    first_data = header_row + 1
    for offset, student in enumerate(students):
        current = first_data + offset
        _id_cell(sheet, current, 1, student["student_id"])
        sheet.cell(row=current, column=2, value=student["name"])

        for index, subject in enumerate(subjects):
            column = 3 + index
            data = student["subjects"].get(subject)
            cell = sheet.cell(row=current, column=column)
            if not data or data["best"] is None:
                cell.value = "—"
                cell.font = Font(name=FONT, size=10, color=MUTED)
                continue
            best: Grade = data["best"]
            target = quote_sheetname(subject_sheets[subject])
            first, last = subject_extent[subject]
            if best.value is not None and student["student_id"] and student["student_id"].isdigit():
                # Puxa a nota final (coluna G, já arredondada) da folha da UC,
                # para que uma correcção lá se reflicta aqui.
                cell.value = (f"=IFERROR(INDEX({target}!$G${first}:$G${last},"
                              f"MATCH($A{current},{target}!$A${first}:$A${last},0)),\"—\")")
                cell.number_format = "0"
            elif best.value is not None:
                cell.value = round_grade(best.value)
                cell.number_format = "0"
            else:
                cell.value = best.label_in(lang)
                cell.font = Font(name=FONT, size=10, italic=True, color=MUTED)

        first_col = get_column_letter(3)
        last_col = get_column_letter(2 + len(subjects))
        span = f"{first_col}{current}:{last_col}{current}"
        average = sheet.cell(row=current, column=3 + len(subjects),
                             value=f'=IFERROR(ROUND(AVERAGE({span}),2),"—")')
        average.number_format = "0.00"
        # Cada UC tem a sua nota minima, por isso a contagem e coluna a coluna.
        counts = "+".join(
            f'COUNTIF({get_column_letter(3 + index)}{current},">="&{pass_refs[subject]})'
            for index, subject in enumerate(subjects)) or "0"
        sheet.cell(row=current, column=4 + len(subjects), value=f"={counts}")
        sheet.cell(row=current, column=5 + len(subjects), value=len(student["subjects"]))

    last_data = first_data + len(students) - 1
    if students:
        _style_body(sheet, first_data, last_data, width)
        # A formatacao usa a nota minima de cada coluna.
        for index, subject in enumerate(subjects):
            letter = get_column_letter(3 + index)
            _grade_rules(sheet, f"{letter}{first_data}:{letter}{last_data}",
                         pass_marks.get(subject, default_pass))
        media = get_column_letter(3 + len(subjects))
        _grade_rules(sheet, f"{media}{first_data}:{media}{last_data}", default_pass)
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{last_data}"

    sheet.freeze_panes = sheet.cell(row=first_data, column=3)

    if students and subjects:
        _subject_footer(sheet, first_data, last_data, subjects, pass_refs, lang)
        _distribution_block(sheet, last_data + 5, subjects, subject_sheets,
                            subject_extent, lang)
    return first_data


def _subject_footer(sheet: Worksheet, first_data: int, last_data: int,
                    subjects: list, pass_refs: dict,
                    lang: str = DEFAULT_LANGUAGE) -> None:
    """Media e aprovados de cada UC, por baixo da coluna dela.

    Somadas entre cadeiras estas contas nao dizem nada; debaixo da coluna de
    cada uma dizem tudo. Sao formulas -- corrigir uma nota actualiza-as -- e
    seguem a nota minima dessa UC, nao um valor global.
    """
    linha_media = last_data + 1
    linha_aprovados = last_data + 2

    for row, chave in ((linha_media, "xl.summary.uc_average"),
                       (linha_aprovados, "xl.summary.uc_approved")):
        etiqueta = sheet.cell(row=row, column=2, value=tr(chave, lang))
        etiqueta.font = Font(name=FONT, size=9.5, bold=True, color=INK)
        etiqueta.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    for index, subject in enumerate(subjects):
        letter = get_column_letter(3 + index)
        intervalo = f"{letter}{first_data}:{letter}{last_data}"

        media = sheet.cell(row=linha_media, column=3 + index,
                           value=f'=IFERROR(ROUND(AVERAGE({intervalo}),2),"—")')
        media.number_format = "0.00"
        media.font = Font(name=FONT, size=10, bold=True, color=INK)
        media.alignment = Alignment(horizontal="center", vertical="center")

        referencia = pass_refs.get(subject)
        aprovados = sheet.cell(row=linha_aprovados, column=3 + index)
        aprovados.value = (f'=COUNTIF({intervalo},">="&{referencia})&"/"&'
                           f'COUNT({intervalo})' if referencia
                           else f"=COUNT({intervalo})")
        aprovados.font = Font(name=FONT, size=9.5, color=MUTED)
        aprovados.alignment = Alignment(horizontal="center", vertical="center")

    borda = Side(style="thin", color=INK_SOFT)
    for column in range(2, 3 + len(subjects)):
        sheet.cell(row=linha_media, column=column).border = Border(top=borda)


def _distribution_block(sheet: Worksheet, row: int, subjects: list,
                        subject_sheets: dict, subject_extent: dict,
                        lang: str = DEFAULT_LANGUAGE) -> None:
    """Distribuicao das melhores notas por escalao, com grafico."""
    bands = [("< 10", -0.001, 9.4999), ("10 – 13", 9.5, 13.4999),
             ("14 – 15", 13.5, 15.4999), ("16 – 17", 15.5, 17.4999),
             ("18 – 20", 17.5, 20.0)]

    sheet.cell(row=row, column=1,
               value=tr("xl.distribution.title", lang)).font = Font(
        name=FONT, size=12, bold=True, color=INK)
    row += 1

    _header_row(sheet, row, [tr("xl.col.band", lang)] + subjects)
    header_row = row
    for index, (label, low, high) in enumerate(bands, start=1):
        current = header_row + index
        sheet.cell(row=current, column=1, value=label)
        for offset, subject in enumerate(subjects):
            target = quote_sheetname(subject_sheets[subject])
            first, last = subject_extent[subject]
            span = f"{target}!$G${first}:$G${last}"
            sheet.cell(row=current, column=2 + offset,
                       value=f'=COUNTIFS({span},">="&{low},{span},"<="&{high})')
    last = header_row + len(bands)
    _style_body(sheet, header_row + 1, last, 1 + len(subjects), centre_from=2)

    chart = BarChart()
    chart.type = "col"
    chart.title = tr("xl.distribution.chart", lang)
    chart.style = 10
    chart.y_axis.title = tr("xl.distribution.students", lang)
    chart.x_axis.title = tr("xl.col.band", lang)
    data = Reference(sheet, min_col=2, min_row=header_row,
                     max_col=1 + len(subjects), max_row=last)
    categories = Reference(sheet, min_col=1, min_row=header_row + 1, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height, chart.width = 8, 16
    sheet.add_chart(chart, f"{get_column_letter(3 + len(subjects))}{header_row}")


# --------------------------------------------------------------------------
# Folha por UC
# --------------------------------------------------------------------------

def _build_subject_sheet(sheet: Worksheet, subject: str, students: list,
                         pass_mark: float, stamp: str,
                         files: Optional[list] = None,
                         lang: str = DEFAULT_LANGUAGE) -> None:
    rows = [s for s in students if subject in s["subjects"]]

    headers = [tr("xl.col.student_id", lang), tr("xl.col.name", lang),
               epoca_label(EPOCAS[0], lang), epoca_label(EPOCAS[1], lang),
               epoca_label(EPOCAS[2], lang), tr("xl.col.best", lang),
               tr("xl.col.final", lang), tr("xl.col.best_epoca", lang),
               tr("xl.col.state", lang), tr("xl.col.best_source", lang)]
    width = len(headers)

    subtitle = tr("xl.subject.subtitle", lang, count=len(rows), stamp=stamp)
    if files:
        subtitle += " · " + tr("xl.subject.sources", lang, files=", ".join(files))
    _title_block(sheet, width, subject, subtitle)

    # Cada cadeira tem a sua nota minima. Fica aqui, editavel, e as colunas
    # «Estado» desta folha e as contagens do Resumo seguem-na.
    sheet.cell(row=SUBJECT_PASS_ROW, column=1,
               value=tr("xl.subject.pass_mark", lang)).font = Font(
        name=FONT, size=10, bold=True)
    pass_cell = sheet.cell(row=SUBJECT_PASS_ROW, column=2, value=pass_mark)
    pass_cell.number_format = GRADE_FORMAT
    pass_cell.font = Font(name=FONT, size=10, bold=True, color="0000FF")
    pass_cell.fill = PatternFill("solid", fgColor="FFF9C4")
    pass_cell.alignment = Alignment(horizontal="center")
    sheet.cell(row=SUBJECT_PASS_ROW, column=3,
               value=tr("xl.subject.pass_hint", lang)).font = Font(
        name=FONT, size=9, italic=True, color=MUTED)

    widths = [11, 40, 13, 13, 15, 13, 12, 16, 13, 30]
    header_row = SUBJECT_HEADER_ROW
    _header_row(sheet, header_row, headers, widths)

    first_data = header_row + 1
    for offset, student in enumerate(rows):
        current = first_data + offset
        data = student["subjects"][subject]
        _id_cell(sheet, current, 1, student["student_id"])
        sheet.cell(row=current, column=2, value=student["name"])

        for index, epoca in enumerate(EPOCAS):
            info = data["epocas"].get(epoca)
            _grade_cell(sheet, current, 3 + index, info["grade"] if info else None, lang)

        span = f"C{current}:E{current}"
        best: Optional[Grade] = data["best"]
        best_cell = sheet.cell(row=current, column=6)
        if data.get("edited") and best is not None and best.value is not None:
            # Nota corrigida a mao: vale mais do que as epocas, por isso vai
            # como valor e nao como a formula que le a melhor delas.
            best_cell.value = best.value
            best_cell.number_format = GRADE_FORMAT
            best_cell.font = Font(name=FONT, size=10, bold=True, color=INK_SOFT)
        elif best is not None and best.value is not None:
            # Recalcula se alguem corrigir uma das epocas.
            best_cell.value = f'=IF(COUNT({span})=0,"—",MAX({span}))'
            best_cell.number_format = GRADE_FORMAT
        elif best is not None:
            best_cell.value = best.label_in(lang)
            best_cell.font = Font(name=FONT, size=10, italic=True, color=MUTED)
        else:
            best_cell.value = "—"

        rounded = sheet.cell(row=current, column=7)
        if best is not None and best.value is not None:
            rounded.value = f'=IF(ISNUMBER(F{current}),ROUND(F{current},0),"—")'
            rounded.number_format = "0"
            rounded.font = Font(name=FONT, size=10, bold=True)
        else:
            rounded.value = "—"

        epoca_cell = sheet.cell(row=current, column=8)
        if data.get("edited"):
            # Uma nota escrita à mão não vem de época nenhuma.
            epoca_cell.value = "—"
        elif best is not None and best.value is not None:
            epoca_cell.value = (f'=IF(COUNT({span})=0,"—",'
                                f'INDEX($C${header_row}:$E${header_row},'
                                f'MATCH(MAX({span}),{span},0)))')
        else:
            epoca_cell.value = (epoca_label(data["best_epoca"], lang)
                                if data.get("best_epoca") else "—")

        aprovado = tr("xl.state.approved", lang)
        reprovado = tr("xl.state.failed", lang)
        state = sheet.cell(row=current, column=9)
        state.value = (f'=IF(NOT(ISNUMBER(F{current})),"—",'
                       f'IF(F{current}>=$B${SUBJECT_PASS_ROW},'
                       f'"{aprovado}","{reprovado}"))'
                       if best is not None and best.value is not None
                       else _state_label(data["approved"], lang))

        best_info = data["epocas"].get(data["best_epoca"] or "")
        origem = sheet.cell(row=current, column=10,
                            value=(tr("xl.edited", lang) if data.get("edited")
                                   else (best_info or {}).get("source_label", "—")))
        if data.get("edited"):
            origem.font = Font(name=FONT, size=9, italic=True, color=INK_SOFT)

    last_data = first_data + len(rows) - 1
    if rows:
        _style_body(sheet, first_data, last_data, width)
        _grade_rules(sheet, f"C{first_data}:G{last_data}", pass_mark)
        _state_rules(sheet, f"I{first_data}:I{last_data}", lang)
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{last_data}"

        for column in range(6, 8):
            for current in range(first_data, last_data + 1):
                sheet.cell(row=current, column=column).font = Font(
                    name=FONT, size=10, bold=True)

    sheet.freeze_panes = sheet.cell(row=first_data, column=3)


def _state_label(approved: Optional[bool], lang: str = DEFAULT_LANGUAGE) -> str:
    if approved is True:
        return tr("xl.state.approved", lang)
    if approved is False:
        return tr("xl.state.failed", lang)
    return tr("xl.state.pending", lang)


def _state_rules(sheet: Worksheet, ref: str, lang: str = DEFAULT_LANGUAGE) -> None:
    sheet.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=[f'"{tr("xl.state.approved", lang)}"'],
        font=Font(name=FONT, size=10, bold=True, color=GREEN),
        fill=PatternFill("solid", bgColor=GREEN_BG)))
    sheet.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=[f'"{tr("xl.state.failed", lang)}"'],
        font=Font(name=FONT, size=10, bold=True, color=RED),
        fill=PatternFill("solid", bgColor=RED_BG)))


# --------------------------------------------------------------------------
# Folha "Médias"
# --------------------------------------------------------------------------

def _build_averages(sheet: Worksheet, students: list, subjects: list,
                    subject_sheets: dict, resumo_first: int, default_pass: float,
                    curriculum: dict, stamp: str, summary_name: str = "Resumo",
                    lang: str = DEFAULT_LANGUAGE) -> None:
    """Medias por semestre, por ano e de fim de curso.

    Cada UC tem aqui duas colunas de apoio -- quanto contribui e quanto pesa --
    e as medias sao a divisao de uma soma pela outra. Assim tudo recalcula: uma
    nota corrigida na folha da UC atravessa o Resumo e chega aqui. Quando os
    ECTS nao estao preenchidos o peso e 1, o que da a media simples.
    """
    resumo = quote_sheetname(summary_name)

    # Uma cadeira sem ano ou semestre nao entra nas medias parciais.
    grupos: dict = {}
    anos: dict = {}
    for subject in subjects:
        meta = curriculum.get(subject) or {}
        year, semester = meta.get("year"), meta.get("semester")
        if year is None or semester is None:
            continue
        grupos.setdefault((year, semester), []).append(subject)
        anos.setdefault(year, []).append(subject)

    headers = [tr("xl.col.student_id", lang), tr("xl.col.name", lang)]
    for subject in subjects:
        headers += [f"{subject} · +", f"{subject} · ×"]
    grupo_keys = sorted(grupos)
    headers += [tr("xl.averages.semester", lang, year=y, semester=s)
                for y, s in grupo_keys]
    ano_keys = sorted(anos)
    headers += [tr("xl.averages.year", lang, year=y) for y in ano_keys]
    headers += [tr("xl.averages.final", lang), tr("xl.col.rounded", lang),
                tr("xl.col.subjects", lang), tr("xl.col.ects", lang)]
    width = len(headers)

    row = _title_block(sheet, width, tr("xl.averages.title", lang),
                       tr("xl.averages.subtitle", lang, stamp=stamp))

    widths = [11, 34] + [14, 10] * len(subjects) + [15] * len(grupo_keys) \
             + [15] * len(ano_keys) + [15, 12, 8, 9]
    header_row = row
    _header_row(sheet, header_row, headers, widths)

    primeira_uc = 3
    coluna_de: dict = {}
    for index, subject in enumerate(subjects):
        coluna_de[subject] = primeira_uc + index * 2

    first_data = header_row + 1
    for offset, student in enumerate(students):
        current = first_data + offset
        origem = resumo_first + offset
        _id_cell(sheet, current, 1, student["student_id"])
        sheet.cell(row=current, column=2, value=student["name"])

        for index, subject in enumerate(subjects):
            meta = curriculum.get(subject) or {}
            ects = meta.get("ects") or 1
            nota = f"{resumo}!{get_column_letter(3 + index)}{origem}"
            minima = (f"{quote_sheetname(subject_sheets[subject])}"
                      f"!$B${SUBJECT_PASS_ROW}")
            conta = f"AND(ISNUMBER({nota}),{nota}>={minima})"
            base = coluna_de[subject]
            sheet.cell(row=current, column=base,
                       value=f"=IF({conta},{nota}*{ects},0)").number_format = "0.##"
            sheet.cell(row=current, column=base + 1,
                       value=f"=IF({conta},{ects},0)").number_format = "0.##"

        coluna = primeira_uc + len(subjects) * 2
        for key in grupo_keys:
            _average_cell(sheet, current, coluna, grupos[key], coluna_de)
            coluna += 1
        for year in ano_keys:
            _average_cell(sheet, current, coluna, anos[year], coluna_de)
            coluna += 1

        _average_cell(sheet, current, coluna, subjects, coluna_de)
        media = f"{get_column_letter(coluna)}{current}"
        sheet.cell(row=current, column=coluna + 1,
                   value=f'=IF(ISNUMBER({media}),ROUND({media},0),"—")').number_format = "0"
        pesos = "+".join(f"IF({get_column_letter(coluna_de[s] + 1)}{current}>0,1,0)"
                         for s in subjects) or "0"
        sheet.cell(row=current, column=coluna + 2, value=f"={pesos}")
        somas = "+".join(f"{get_column_letter(coluna_de[s] + 1)}{current}"
                         for s in subjects) or "0"
        sheet.cell(row=current, column=coluna + 3, value=f"={somas}").number_format = "0.##"

    last_data = first_data + len(students) - 1
    if students:
        _style_body(sheet, first_data, last_data, width)
        inicio = get_column_letter(primeira_uc + len(subjects) * 2)
        fim = get_column_letter(primeira_uc + len(subjects) * 2
                                + len(grupo_keys) + len(ano_keys))
        _grade_rules(sheet, f"{inicio}{first_data}:{fim}{last_data}", default_pass)
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{last_data}"
        if subjects:
            # As colunas de apoio ficam agrupadas: dao para recolher.
            sheet.column_dimensions.group(
                get_column_letter(primeira_uc),
                get_column_letter(primeira_uc + len(subjects) * 2 - 1), hidden=True)

    sheet.freeze_panes = sheet.cell(row=first_data, column=3)


def _average_cell(sheet: Worksheet, row: int, column: int, subjects: list,
                  coluna_de: dict) -> None:
    """Soma das contribuicoes a dividir pela soma dos pesos."""
    if not subjects:
        sheet.cell(row=row, column=column, value="—")
        return
    contribui = "+".join(f"{get_column_letter(coluna_de[s])}{row}" for s in subjects)
    peso = "+".join(f"{get_column_letter(coluna_de[s] + 1)}{row}" for s in subjects)
    cell = sheet.cell(row=row, column=column,
                      value=f'=IF(({peso})=0,"—",ROUND(({contribui})/({peso}),2))')
    cell.number_format = "0.00"
    cell.font = Font(name=FONT, size=10, bold=True)


# --------------------------------------------------------------------------
# Folha "Detalhe"
# --------------------------------------------------------------------------

def _build_detail(sheet: Worksheet, students: list, subjects: list, stamp: str,
                  lang: str = DEFAULT_LANGUAGE) -> None:
    headers = [tr("xl.col.student_id", lang), tr("xl.col.name", lang),
               tr("xl.col.subject", lang), tr("xl.col.epoca", lang),
               tr("xl.col.kind", lang), tr("xl.col.item", lang),
               tr("xl.col.value", lang), tr("xl.col.source", lang)]
    width = len(headers)
    row = _title_block(sheet, width, tr("xl.detail.title", lang),
                       tr("xl.detail.subtitle", lang, stamp=stamp))
    _header_row(sheet, row, headers, [11, 34, 32, 15, 13, 22, 11, 34])
    header_row = row
    current = header_row + 1

    for student in students:
        for subject in subjects:
            data = student["subjects"].get(subject)
            if not data:
                continue
            for epoca in EPOCAS:
                info = data["epocas"].get(epoca)
                if not info:
                    continue
                _id_cell(sheet, current, 1, student["student_id"])
                sheet.cell(row=current, column=2, value=student["name"])
                sheet.cell(row=current, column=3, value=subject)
                sheet.cell(row=current, column=4, value=epoca_label(epoca, lang))
                sheet.cell(row=current, column=5, value=tr("xl.detail.final", lang))
                sheet.cell(row=current, column=6, value=info["column"] or "—")
                _grade_cell(sheet, current, 7, info["grade"], lang)
                sheet.cell(row=current, column=8, value=info["source_label"])
                current += 1

    last = current - 1
    if last >= header_row + 1:
        _style_body(sheet, header_row + 1, last, width)
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{last}"
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


# --------------------------------------------------------------------------
# Folha "Avisos"
# --------------------------------------------------------------------------

def _build_notices(sheet: Worksheet, result: dict, stamp: str,
                   lang: str = DEFAULT_LANGUAGE) -> None:
    headers = [tr("xl.col.severity", lang), tr("xl.col.type", lang),
               tr("xl.col.name", lang), tr("xl.col.subject", lang),
               tr("xl.col.description", lang), tr("xl.col.chosen", lang)]
    width = len(headers)
    entries = list(result.get("conflicts", [])) + list(result.get("warnings", []))

    row = _title_block(sheet, width, tr("xl.notices.title", lang),
                       tr("xl.notices.subtitle", lang, count=len(entries), stamp=stamp))
    _header_row(sheet, row, headers, [13, 18, 32, 30, 78, 32])
    header_row = row
    current = header_row + 1

    if not entries:
        sheet.cell(row=current, column=1, value=tr("xl.notices.none", lang))
        sheet.cell(row=current, column=1).font = Font(name=FONT, size=10, italic=True,
                                                      color=GREEN)
        return

    for entry in entries:
        severity = entry.get("severity", "info")
        cell = sheet.cell(row=current, column=1,
                          value=tr("xl.notices.conflict" if severity == "warning"
                                   else "xl.notices.info", lang))
        cell.font = Font(name=FONT, size=10, bold=True,
                         color=AMBER if severity == "warning" else INK_SOFT)
        sheet.cell(row=current, column=2, value=notice_type(entry.get("type"), lang))
        sheet.cell(row=current, column=3, value=entry.get("student", ""))
        sheet.cell(row=current, column=4, value=entry.get("subject", ""))
        sheet.cell(row=current, column=5, value=render(entry.get("detail", ""), lang))
        sheet.cell(row=current, column=6, value=render(entry.get("chosen", ""), lang))
        current += 1

    last = current - 1
    _style_body(sheet, header_row + 1, last, width, centre_from=99)
    for r in range(header_row + 1, last + 1):
        sheet.cell(row=r, column=5).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True, indent=1)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{last}"
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
